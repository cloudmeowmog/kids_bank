# -*- coding: utf-8 -*-
"""
孩子生活管理存款系統 — Streamlit 版（Google Sheets 雲端儲存）
原桌面版 (tkinter) 改寫，可部署於 GitHub + Streamlit Community Cloud，手機瀏覽器可直接使用。

【資料儲存】
- 若在 st.secrets 設好 Google 服務帳號與試算表 → 自動把資料存到 Google Sheets（雲端持久保存，重開不遺失）。
- 若未設定 secrets（例如本機測試）→ 退回使用本機 JSON 檔 kids_bank_data.json。
設定方式請見 README.md。資料模型不變，整包 JSON 直接寫進試算表（超過單格上限會自動分段）。
"""

import json
import os
from datetime import datetime, timedelta
from io import BytesIO

import streamlit as st

# ---------------------------------------------------------------------------
# 設定
# ---------------------------------------------------------------------------
DATA_FILE = "kids_bank_data.json"          # 本機後援用檔名
GSHEET_WORKSHEET = "data"                  # 存放 JSON 的工作表名稱
CHUNK_SIZE = 45000                         # 單格字數上限 5 萬，保守取 4.5 萬分段

REWARD_OPTIONS = {
    "幫忙清潔廁所": 50,
    "主動幫忙洗碗": 10,
    "主動幫忙倒垃圾": 10,
    "寢室保持清潔": 50,
    "寢室衣物收拾整齊": 50,
    "主動協助收衣服": 10,
    "主動協助晾衣服": 20,
    "其他": None,
}

PUNISHMENT_OPTIONS = {
    "吃完東西未收拾": 10,
    "作業未於規定時間完成": 50,
    "個人物品隨意擺放": 10,
    "態度不佳": 50,
    "未經允許使用3C產品": 100,
    "3C產品使用超過規範時間": 50,
    "其他": None,
}


# ---------------------------------------------------------------------------
# 資料存取（純函式，方便測試）
# ---------------------------------------------------------------------------
def _default_data(today_str=None):
    if today_str is None:
        today_str = datetime.now().strftime("%Y-%m-%d")
    return {
        "users": {
            "Luke": {"balance": 0.0, "rate": 0.01, "open_date": today_str, "history": []},
            "Leia": {"balance": 0.0, "rate": 0.01, "open_date": today_str, "history": []},
        },
        "last_update": today_str,
        "daily_reward": 50.0,
    }


# ---- 後端選擇：有設定 Google 服務帳號就用 Sheets，否則用本機檔 ----
def _use_gsheets():
    try:
        return ("gcp_service_account" in st.secrets) and ("spreadsheet" in st.secrets)
    except Exception:
        return False


@st.cache_resource(show_spinner=False)
def _get_worksheet():
    """建立並快取 gspread 連線與工作表（避免每次 rerun 重新驗證）。"""
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    info = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(info, scopes=scopes)
    gc = gspread.authorize(creds)

    conf = st.secrets["spreadsheet"]
    if "key" in conf:
        sh = gc.open_by_key(conf["key"])
    elif "url" in conf:
        sh = gc.open_by_url(conf["url"])
    else:
        raise RuntimeError("st.secrets['spreadsheet'] 需要設定 key 或 url")

    try:
        ws = sh.worksheet(GSHEET_WORKSHEET)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=GSHEET_WORKSHEET, rows=200, cols=2)
    return ws


def _gs_load():
    ws = _get_worksheet()
    blob = "".join(ws.col_values(1)).strip()   # A 欄各格串接還原
    if not blob:
        data = _default_data()
        _gs_save(data)
        return data
    try:
        data = json.loads(blob)
        if "users" not in data or "last_update" not in data:
            raise ValueError("資料格式不正確")
        return data
    except (json.JSONDecodeError, ValueError):
        # 損壞：把壞掉的內容備份到 B1，再重建預設
        try:
            ws.update(range_name="B1", values=[[blob[:45000]]], raw=True)
        except Exception:
            pass
        data = _default_data()
        _gs_save(data)
        return data


def _gs_save(data):
    ws = _get_worksheet()
    blob = json.dumps(data, ensure_ascii=False)
    chunks = [blob[i:i + CHUNK_SIZE] for i in range(0, len(blob), CHUNK_SIZE)] or [""]
    ws.clear()
    ws.update(range_name=f"A1:A{len(chunks)}", values=[[c] for c in chunks], raw=True)


def _file_load():
    """本機後援：讀取 JSON 檔；不存在則建立；損壞則備份後重建。"""
    if not os.path.exists(DATA_FILE):
        data = _default_data()
        _file_save(data)
        return data
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if "users" not in data or "last_update" not in data:
            raise ValueError("資料檔格式不正確")
        return data
    except (json.JSONDecodeError, ValueError):
        if os.path.exists(DATA_FILE):
            try:
                os.replace(DATA_FILE, DATA_FILE + ".bak")
            except OSError:
                pass
        data = _default_data()
        _file_save(data)
        return data


def _file_save(data):
    """原子寫入：先寫暫存檔再 replace，避免寫到一半被中斷造成 JSON 損壞。"""
    tmp = DATA_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    os.replace(tmp, DATA_FILE)


def load_data():
    return _gs_load() if _use_gsheets() else _file_load()


def save_data(data):
    if _use_gsheets():
        _gs_save(data)
    else:
        _file_save(data)


# ---------------------------------------------------------------------------
# 自動補發每日獎勵 / 每月利息
# ---------------------------------------------------------------------------
def auto_update_records(data):
    """從開戶日期起，補上至今缺少的每日獎勵與每月利息。
    - 已存在的獎勵/利息日期不會被變更
    - 利息於每月 1 日 08:00 結算，以當時帳戶餘額計算
    - 每日獎勵於 22:30 發放（已過 22:30 才含今天）
    回傳 True 代表有補發並已存檔。
    """
    today = datetime.now().date()
    now = datetime.now()
    daily_reward = data.get("daily_reward", 0)
    updated = False

    for child_name, info in data["users"].items():
        existing_reward_dates = set()
        for r in info["history"]:
            if r["type"] == "系統-每日獎勵":
                existing_reward_dates.add(r["date"][:10])

        existing_interest_months = set()
        for r in info["history"]:
            if r["type"] == "系統-每月利息":
                existing_interest_months.add(r["date"][:7])

        open_date_str = info.get("open_date", data.get("last_update", today.strftime("%Y-%m-%d")))
        open_date = datetime.strptime(open_date_str, "%Y-%m-%d").date()
        start_date = open_date

        # 已過 22:30 則含今天，否則到昨天
        if now.hour > 22 or (now.hour == 22 and now.minute >= 30):
            end_date = today
        else:
            end_date = today - timedelta(days=1)

        new_records = []
        current_date = start_date
        while current_date <= end_date:
            date_key = current_date.strftime("%Y-%m-%d")
            month_key = current_date.strftime("%Y-%m")

            if current_date.day == 1 and month_key not in existing_interest_months:
                rate_pct = f"{info['rate'] * 100:.2f}".rstrip("0").rstrip(".")
                new_records.append({
                    "date": date_key + " 08:00",
                    "type": "系統-每月利息",
                    "amount": None,
                    "note": f"利率 {rate_pct}%",
                    "_is_interest": True,
                })

            if daily_reward > 0 and date_key not in existing_reward_dates:
                new_records.append({
                    "date": date_key + " 22:30",
                    "type": "系統-每日獎勵",
                    "amount": daily_reward,
                    "note": "每日固定配給 (22:30發放)",
                })

            current_date += timedelta(days=1)

        if not new_records:
            continue

        for rec in new_records:
            if rec.get("_is_interest"):
                rec["amount"] = 0
                rec["balance"] = 0
                del rec["_is_interest"]
            else:
                rec["balance"] = 0
            info["history"].append(rec)

        info["history"].sort(key=lambda r: r["date"])

        current_balance = 0.0
        for record in info["history"]:
            if record["type"] == "系統-每月利息" and record["amount"] == 0:
                interest = round(current_balance * info["rate"])
                if interest > 0:
                    record["amount"] = interest
                else:
                    record["_remove"] = True
            current_balance += record["amount"]
            record["balance"] = round(current_balance)

        info["history"] = [r for r in info["history"] if not r.get("_remove")]

        current_balance = 0.0
        for record in info["history"]:
            current_balance += record["amount"]
            record["balance"] = round(current_balance)
        info["balance"] = round(current_balance)

        updated = True

    if updated:
        data["last_update"] = today.strftime("%Y-%m-%d")
        save_data(data)
        return True
    return False


def recalculate_balances(data, child_name):
    user = data["users"][child_name]
    current_balance = 0.0
    for record in user["history"]:
        current_balance += record["amount"]
        record["balance"] = round(current_balance)
    user["balance"] = round(current_balance)


def fmt_rate(rate):
    return f"{rate * 100:.2f}".rstrip("0").rstrip(".")


# ---------------------------------------------------------------------------
# Excel 匯出 / 匯入
# ---------------------------------------------------------------------------
def build_excel_bytes(target, user, selected_month):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{target} 存款簿"

    header_font = Font(name="微軟正黑體", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FF9800")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{target} 的存款簿"
    ws["A1"].font = Font(name="微軟正黑體", bold=True, size=16, color="4E342E")
    ws["A1"].alignment = Alignment(horizontal="center")

    rate_pct = fmt_rate(user["rate"])
    ws.merge_cells("A2:E2")
    ws["A2"] = (f"目前餘額: {int(round(user['balance']))} 元 | 利率: {rate_pct}% | "
                f"匯出時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws["A2"].font = Font(name="微軟正黑體", size=10, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    if selected_month != "全部":
        ws.merge_cells("A3:E3")
        ws["A3"] = f"篩選月份: {selected_month}"
        ws["A3"].font = Font(name="微軟正黑體", size=10, color="D84315")
        ws["A3"].alignment = Alignment(horizontal="center")
        header_row = 5
    else:
        header_row = 4

    headers = ["日期時間", "類型", "異動金額", "目前結餘", "備註說明"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name="微軟正黑體", size=11)
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_left = Alignment(horizontal="left", vertical="center")

    row_idx = header_row + 1
    for record in user["history"]:
        if selected_month != "全部" and not record["date"].startswith(selected_month):
            continue
        amt = int(round(record["amount"]))
        bal = int(round(record["balance"]))
        amt_str = f"+{amt}" if amt > 0 else str(amt)

        row_data = [record["date"], record["type"], amt_str, bal, record.get("note", "")]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_align_center if col <= 4 else data_align_left

        amt_cell = ws.cell(row=row_idx, column=3)
        if amt > 0:
            amt_cell.font = Font(name="微軟正黑體", size=11, color="2E7D32")
        elif amt < 0:
            amt_cell.font = Font(name="微軟正黑體", size=11, color="D32F2F")

        row_idx += 1

    for col, width in zip("ABCDE", [20, 16, 14, 14, 30]):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_excel_records(file_bytes):
    """從上傳的 Excel bytes 解析出紀錄列表。回傳 (records, error_msg)。"""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return None, f"無法讀取 Excel：{e}"
    ws = wb.active

    header_row = None
    for row in range(1, min(ws.max_row + 1, 10)):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and "日期" in str(cell_val):
            header_row = row
            break
    if header_row is None:
        return None, "找不到表頭列（需包含「日期時間」欄位）。請確認格式與匯出格式一致。"

    imported = []
    for row in range(header_row + 1, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        type_val = ws.cell(row=row, column=2).value
        amt_val = ws.cell(row=row, column=3).value
        note_val = ws.cell(row=row, column=5).value or ""

        if not date_val or not type_val:
            continue

        amt_str = str(amt_val).replace("+", "").strip()
        try:
            amount = float(amt_str)
        except (ValueError, TypeError):
            continue

        imported.append({
            "date": str(date_val).strip(),
            "type": str(type_val).strip(),
            "amount": amount,
            "balance": 0,
            "note": str(note_val).strip(),
        })
    return imported, None


# ===========================================================================
# Streamlit UI
# ===========================================================================
st.set_page_config(page_title="孩子存款系統", page_icon="🏦", layout="centered")

# --- 暖色系外觀 ---
st.markdown(
    """
    <style>
      .stApp { background-color: #FFF8E1; }
      h1, h2, h3, h4, label, p, .stMarkdown { color: #4E342E !important; }
      div.stButton > button {
          background-color: #FFCC80; color: #4E342E; font-weight: 700;
          border: 0; border-radius: 10px; padding: 0.45rem 1rem;
      }
      div.stButton > button:hover { background-color: #FFA726; color: #4E342E; }
      div[data-testid="stMetricValue"] { color: #D84315 !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

# --- 載入資料（每次 rerun 都讀最新檔），並自動補發 ---
data = load_data()
if auto_update_records(data):
    st.toast("已自動補齊漏掉的每日獎勵與利息結算！", icon="🎉")

children = list(data["users"].keys())

st.title("🏦 孩子生活管理存款系統")

# --- 側邊欄：餘額總覽 + 備份/還原 ---
with st.sidebar:
    if _use_gsheets():
        st.success("☁️ 儲存：Google Sheets", icon="✅")
    else:
        st.warning("💾 儲存：本機檔案（雲端會遺失，請設定 Google Sheets）", icon="⚠️")

    st.header("💰 帳戶總覽")
    for c in children:
        u = data["users"][c]
        st.metric(f"{c}（利率 {fmt_rate(u['rate'])}%）", f"{int(round(u['balance']))} 元")

    st.divider()
    st.subheader("🗄️ 資料備份 / 還原")
    st.caption("雲端部署檔案不會長久保存，請定期下載備份。")

    st.download_button(
        "⬇️ 下載備份 (JSON)",
        data=json.dumps(data, ensure_ascii=False, indent=4).encode("utf-8"),
        file_name=f"kids_bank_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
        mime="application/json",
        use_container_width=True,
    )

    restore_file = st.file_uploader("⬆️ 上傳備份還原", type=["json"], key="restore")
    if restore_file is not None:
        try:
            restored = json.loads(restore_file.read().decode("utf-8"))
            if "users" not in restored or "last_update" not in restored:
                raise ValueError("格式不正確")
            if st.button("確認還原（覆蓋目前資料）", type="primary", use_container_width=True):
                save_data(restored)
                st.success("已還原！")
                st.rerun()
        except Exception as e:
            st.error(f"還原失敗：{e}")

# --- 主分頁 ---
tab_trans, tab_passbook, tab_opendate, tab_settings = st.tabs(
    ["➕ 新增交易", "📒 查看存款簿", "📅 開戶日期", "⚙️ 系統設定"]
)

# ---------------------------------------------------------------------------
# 分頁一：新增交易
# ---------------------------------------------------------------------------
with tab_trans:
    st.subheader("新增交易")

    target = st.selectbox("選擇對象", children + ["兩人同時"], key="tx_target")
    ttype = st.selectbox(
        "交易類型",
        ["額外獎勵 (存入)", "懲罰扣款 (扣除)", "帳戶提取 (領出)"],
        key="tx_type",
    )

    note = ""
    preset_amt = None
    reason = ""

    if ttype == "額外獎勵 (存入)":
        reason = st.selectbox("獎勵原因", list(REWARD_OPTIONS.keys()), key="tx_reward")
        if reason == "其他":
            note = st.text_input("自訂獎勵原因", key="tx_reward_custom").strip()
        else:
            note = reason
            preset_amt = REWARD_OPTIONS[reason]
    elif ttype == "懲罰扣款 (扣除)":
        reason = st.selectbox("扣款原因", list(PUNISHMENT_OPTIONS.keys()), key="tx_pun")
        if reason == "其他":
            note = st.text_input("自訂扣款原因", key="tx_pun_custom").strip()
        else:
            note = reason
            preset_amt = PUNISHMENT_OPTIONS[reason]
    else:  # 帳戶提取
        note = st.text_input("備註 (選填)", key="tx_note").strip()

    # 預設金額：切換項目時自動帶入，但允許使用者再修改
    if "tx_amount" not in st.session_state:
        st.session_state["tx_amount"] = 0.0
    sig = f"{ttype}|{reason}"
    if st.session_state.get("_tx_sig") != sig:
        st.session_state["_tx_sig"] = sig
        st.session_state["tx_amount"] = float(preset_amt) if preset_amt is not None else 0.0

    amount = st.number_input("金額", min_value=0.0, step=1.0, key="tx_amount")

    if st.button("✅ 送出紀錄", use_container_width=True):
        # 驗證
        if ttype == "額外獎勵 (存入)" and not note:
            st.error("請選擇或輸入獎勵原因！")
        elif ttype == "懲罰扣款 (扣除)" and not note:
            st.error("請選擇或輸入扣款原因！")
        elif amount <= 0:
            st.error("金額請輸入大於 0 的數字！")
        else:
            t_type = ttype.split(" ")[0]
            selected_children = children if target == "兩人同時" else [target]
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            msgs = []
            for child in selected_children:
                user = data["users"][child]
                act_amount = amount
                if t_type in ["懲罰扣款", "帳戶提取"]:
                    act_amount = -abs(amount)
                    if user["balance"] + act_amount < 0:
                        msgs.append(f"⚠️ {child} 餘額不足以扣除！(仍允許記錄)")
                user["balance"] += act_amount
                user["history"].append({
                    "date": now_str, "type": t_type, "amount": act_amount,
                    "balance": round(user["balance"]), "note": note,
                })
                msgs.append(
                    f"✅ {child} 的 {t_type} 已記錄：{int(round(act_amount))} 元 "
                    f"(目前餘額 {int(round(user['balance']))} 元)"
                )
            save_data(data)
            st.success("\n\n".join(msgs))
            # 清掉金額，保留選項
            st.session_state.pop("_tx_sig", None)
            st.session_state["tx_amount"] = 0.0
            st.rerun()

# ---------------------------------------------------------------------------
# 分頁二：查看存款簿
# ---------------------------------------------------------------------------
with tab_passbook:
    st.subheader("查看存款簿")

    pb_target = st.selectbox("檢視帳戶", children, key="pb_target")
    user = data["users"][pb_target]

    months = sorted({r["date"][:7] for r in user["history"]}, reverse=True)
    month_list = ["全部"] + months
    pb_month = st.selectbox("月份", month_list, key="pb_month")

    st.metric(
        f"{pb_target} 目前餘額（利率 {fmt_rate(user['rate'])}%）",
        f"{int(round(user['balance']))} 元",
    )

    # 篩選紀錄（保留原始 index）
    rows = []
    for i, r in enumerate(user["history"]):
        if pb_month != "全部" and not r["date"].startswith(pb_month):
            continue
        amt = int(round(r["amount"]))
        rows.append({
            "編號": i,
            "日期時間": r["date"],
            "類型": r["type"],
            "異動金額": f"+{amt}" if amt > 0 else str(amt),
            "目前結餘": int(round(r["balance"])),
            "備註說明": r.get("note", ""),
        })
    rows_desc = list(reversed(rows))

    # 匯出 Excel
    col1, col2 = st.columns(2)
    with col1:
        default_name = f"{pb_target}_存款簿"
        if pb_month != "全部":
            default_name += f"_{pb_month}"
        st.download_button(
            "📥 匯出 Excel",
            data=build_excel_bytes(pb_target, user, pb_month),
            file_name=default_name + ".xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    if rows_desc:
        st.dataframe(rows_desc, use_container_width=True, hide_index=True)
    else:
        st.info("此條件下沒有紀錄。")

    # --- 匯入 Excel ---
    with st.expander("📤 匯入 Excel"):
        up = st.file_uploader("選擇 Excel 檔", type=["xlsx"], key="imp_excel")
        imp_mode = st.radio("匯入模式", ["附加到現有紀錄後面", "取代所有現有紀錄"], key="imp_mode")
        if up is not None and st.button("開始匯入", key="do_import"):
            imported, err = parse_excel_records(up.read())
            if err:
                st.error(err)
            elif not imported:
                st.warning("Excel 中沒有找到可匯入的紀錄。")
            else:
                if imp_mode.startswith("附加"):
                    existing_keys = {(r["date"], r["type"]) for r in user["history"]}
                    new_records = [r for r in imported if (r["date"], r["type"]) not in existing_keys]
                    user["history"].extend(new_records)
                    added, skipped = len(new_records), len(imported) - len(new_records)
                else:
                    user["history"] = imported
                    added, skipped = len(imported), 0

                user["history"].sort(key=lambda r: r["date"])
                recalculate_balances(data, pb_target)
                save_data(data)
                msg = f"匯入完成！新增 {added} 筆。"
                if skipped:
                    msg += f"（跳過 {skipped} 筆重複）"
                st.success(msg)
                st.rerun()

    # --- 編輯 / 刪除 ---
    with st.expander("✏️ 編輯 / 刪除紀錄"):
        if not rows:
            st.caption("目前沒有可編輯的紀錄。")
        else:
            options = {
                f"#{r['編號']}｜{r['日期時間']}｜{r['類型']}｜{r['異動金額']}": r["編號"]
                for r in rows_desc
            }
            picked = st.selectbox("選擇一筆紀錄", list(options.keys()), key="edit_pick")
            idx = options[picked]
            rec = user["history"][idx]

            new_amt = st.number_input(
                "修改金額（輸入正數即可，系統依類型自動帶正負號）",
                min_value=0.0, step=1.0,
                value=float(abs(int(round(rec["amount"])))),
                key="edit_amt",
            )
            new_note = st.text_input("修改備註", value=rec.get("note", ""), key="edit_note")

            ec1, ec2 = st.columns(2)
            with ec1:
                if st.button("💾 儲存修改", use_container_width=True):
                    if "扣" in rec["type"] or "提取" in rec["type"]:
                        rec["amount"] = -abs(new_amt)
                    else:
                        rec["amount"] = abs(new_amt)
                    rec["note"] = new_note
                    recalculate_balances(data, pb_target)
                    save_data(data)
                    st.success("已修改，結餘已更新！")
                    st.rerun()
            with ec2:
                if st.button("🗑️ 刪除此筆", use_container_width=True):
                    del user["history"][idx]
                    recalculate_balances(data, pb_target)
                    save_data(data)
                    st.success("已刪除，結餘已更新！")
                    st.rerun()

            st.divider()
            multi = st.multiselect("批次刪除（可多選）", list(options.keys()), key="multi_del")
            if multi and st.button("🗑️ 刪除選取的多筆", use_container_width=True):
                del_idx = sorted({options[k] for k in multi}, reverse=True)
                for i in del_idx:
                    del user["history"][i]
                recalculate_balances(data, pb_target)
                save_data(data)
                st.success(f"已刪除 {len(del_idx)} 筆，結餘已更新！")
                st.rerun()

# ---------------------------------------------------------------------------
# 分頁三：開戶日期
# ---------------------------------------------------------------------------
with tab_opendate:
    st.subheader("設定各帳戶開戶日期")
    st.caption("程式會從開戶日期起計算每日獎勵。")

    new_dates = {}
    for child in children:
        cur = data["users"][child].get("open_date", datetime.now().strftime("%Y-%m-%d"))
        try:
            cur_d = datetime.strptime(cur, "%Y-%m-%d").date()
        except ValueError:
            cur_d = datetime.now().date()
        new_dates[child] = st.date_input(f"{child} 開戶日期", value=cur_d, key=f"open_{child}")

    oc1, oc2 = st.columns(2)
    with oc1:
        if st.button("💾 儲存開戶日期", use_container_width=True):
            for child, d in new_dates.items():
                data["users"][child]["open_date"] = d.strftime("%Y-%m-%d")
            save_data(data)
            st.success("開戶日期已更新！按「重新計算帳戶」即可依新日期補發。")
    with oc2:
        if st.button("🔄 重新計算帳戶", use_container_width=True):
            for child, d in new_dates.items():
                data["users"][child]["open_date"] = d.strftime("%Y-%m-%d")
            save_data(data)
            if auto_update_records(data):
                st.success("帳戶已重新計算，獎勵與利息已補齊！")
                st.rerun()
            else:
                st.info("所有帳戶皆已完整，無需補發。")

# ---------------------------------------------------------------------------
# 分頁四：系統設定
# ---------------------------------------------------------------------------
with tab_settings:
    st.subheader("系統設定")

    daily = st.number_input(
        "每日固定發放獎勵（元）",
        min_value=0.0, step=1.0,
        value=float(data.get("daily_reward", 0)),
        key="set_daily",
    )

    rate_inputs = {}
    for child in children:
        rate_inputs[child] = st.number_input(
            f"{child} 存款利率（%）",
            min_value=0.0, step=0.01, format="%.2f",
            value=float(f"{data['users'][child]['rate'] * 100:.2f}"),
            key=f"set_rate_{child}",
        )

    if st.button("💾 儲存系統設定", use_container_width=True):
        data["daily_reward"] = daily
        for child, val in rate_inputs.items():
            data["users"][child]["rate"] = round(val / 100, 4)
        save_data(data)
        st.success("參數已儲存！（利率調整不會變更已登錄的舊利息）")
        st.rerun()
